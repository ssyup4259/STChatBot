# -*- coding: utf-8 -*-
import json
import re
import requests
import urllib.request
import urllib.parse
import openpyxl
import googlemaps

from threading import Thread
from bs4 import BeautifulSoup
from flask import Flask, request, session
from slack import WebClient
from slack.web.classes import extract_json
from slack.web.classes.blocks import *
from slack.web.classes.elements import *
from slack.web.classes.interactions import MessageInteractiveEvent
from slackeventsapi import SlackEventAdapter
from random import shuffle



SLACK_TOKEN = "?"
SLACK_SIGNING_SECRET = "?"
app = Flask(__name__)
slack_events_adaptor = SlackEventAdapter(SLACK_SIGNING_SECRET, "/listening", app)
slack_web_client = WebClient(token=SLACK_TOKEN)
# app.secret_key = b'1234qwer'
# xlm 읽어서 tours에 담아 놓기
filename = "seoul_tour.xlsx"
book = openpyxl.load_workbook(filename)
sheet = book.worksheets[0]
sheet = book.active

# 구글 맵 키
gmaps_key = '?'
gmaps = googlemaps.Client(key=gmaps_key)

gRain =""

tours = []
for i in range (2, sheet.max_row) :
    try :
        tour = {
            'Category' : sheet.cell(row=i, column=1 ).value,
            'Primary_num' : sheet.cell(row=i, column=2 ).value,
            'Trade_Name' : sheet.cell(row=i, column=3 ).value,
            'Old_address' : sheet.cell(row=i, column=4 ).value,
            'New_address' : sheet.cell(row=i, column=5 ).value,
            'Dial' : sheet.cell(row=i, column=6 ).value,
            'Tag' : sheet.cell(row=i, column=7 ).value,
            'Dis' : sheet.cell(row=i, column=5 ).value.split()[2]
        }
    except:
        pass
    tours.append(tour)
shuffle(tours)

def inputText(channel, text):
    global gRain
    if "hi" in text:
        block1 = ImageBlock(
            image_url="https://www.seoul.go.kr/res_newseoul/images/seoul/img_intro1.png",
            alt_text="이미지 왜 안나와"
        )
        block2 = SectionBlock(
            text="*안녕하세요. 서울의 관광지를 추천해주는 챗봇입니다.*\n여행 날짜를 골라주세요"
        )
        block3 = ActionsBlock(
            elements=[
                ButtonElement(
                    text="TODAY(오늘)",
                    action_id="1", value=str(0)
                ),
                ButtonElement(
                    text="TOMORROW(내일)", style="danger",
                    action_id="2", value=str(1)
                ),
                ButtonElement(
                    text="AFTER TOMORROW(모레)", style="primary",
                    action_id="3", value=str(2)
                ),
                ButtonElement(
                    text="ETC(다른 날짜)",
                    action_id="4", value=str(3)
                ),
            ]
        )
        my_blocks = [block1, block2, block3]
        slack_web_client.chat_postMessage(
            channel=channel,
            blocks=extract_json(my_blocks)
        )
    elif text == "else":
        slack_web_client.chat_postMessage(
            channel=channel,
            text="아직 기능 구현중입니다. 조금만 기다려주세요"
        )
    else:  # 강남구 같은거 들어오면, 위에 애들 처리 안하고
        index = 1
        flag = False
        print(gRain)
        for titem in tours:
            tag = titem['Dis']
            txt = text.split()[1]
            if index >3 :
                break;
            if txt in tag:
                flag = True
                if gRain == "raining":
                    if titem['Category'] == "명소":
                        continue
                    else:
                        block1 = ImageBlock(
                            image_url="https://is3-ssl.mzstatic.com/image/thumb/Purple114/v4/25/93/85/2593857c-20bc-9ac2-6c96-05362162b745/source/512x512bb.jpg",
                            alt_text="실내"
                        )
                        index += 1
                else:
                    if titem['Category'] != "명소":
                        block1 = ImageBlock(
                            image_url="https://is3-ssl.mzstatic.com/image/thumb/Purple114/v4/25/93/85/2593857c-20bc-9ac2-6c96-05362162b745/source/512x512bb.jpg",
                            alt_text="실내"
                        )
                        index += 1
                    else:
                        block1 = ImageBlock(
                            image_url="http://www.urbanbrush.net/web/wp-content/uploads/edd/2018/05/web-20180531130941307931.png",
                            alt_text="실외"
                        )
                        index += 1
                block2 = SectionBlock(
                    accessory=block1,
                    text="*" + titem['Category'] + "이에요*\n 상호명: " + titem['Trade_Name'] + " 주소: " + titem['Old_address']
                )

                tmp = gmaps.geocode(titem['New_address'])
                at = tmp[0]['geometry']['location']['lat']
                ng = tmp[0]['geometry']['location']['lng']

                block4 = ImageBlock(
                    image_url='https://maps.googleapis.com/maps/api/staticmap?&zoom=15&size=600x300&maptype=roadmap&markers=color:blue%7Clabel:S%7C' + str(
                        at) + ',' + str(ng) +
                              '&markers=color:green%7Clabel:G%7C' + str(at) + ',' + str(ng) +
                              '&markers=color:red%7Clabel:C%7C' + str(at) + ',' + str(ng) +
                              '&key=AIzaSyAVES7wI90_RMlHaDjhceIWVRslqj413kM',

                    alt_text="실외"
                )

                my_blocks = [block2, block4]
                slack_web_client.chat_postMessage(
                    channel=channel,
                    blocks=extract_json(my_blocks)
                )

        if flag == False:
            slack_web_client.chat_postMessage(
                channel=channel,
                text="서울시 내에 존재하는 구단위를 입력해주세요 오타 없이요 제발"
            )

def buttonText(click_event):
    global gRain
    location = '역삼동'
    enc_location = urllib.parse.quote(location + '+날씨')
    url = 'https://search.naver.com/search.naver?ie=utf8&query=' + enc_location
    source_code = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(source_code, "html.parser")

    message = []
    for ul_tag in soup.find_all("ul", class_="_pageList"):
        for li_tag in ul_tag.find_all("li", class_="today"):
            day = li_tag.find("span", class_="day_info").get_text().strip()
            temps = li_tag.find("dd").get_text().strip().replace('°', '').split("/")
            temp = (int(temps[0]) + int(temps[1])) / 2
            rain = li_tag.find("span", class_="morning").find("span", class_="num").get_text().strip()
            weather = day + "의 날씨는 평균 " + str(temp) + "°C 이고 강수확률은 " + rain + "% 입니다"
            if int(rain) > 50:
                weather += "\n비가 올 확률이 높으니 실내 활동을 추천드립니다."
            else:
                weather += "\n날씨가 좋아 야외 활동에 적합합니다."
            message.append(weather)

    txt = ""
    if int(click_event.value) == 3:
        for i in range(3, len(message), 1):
            txt += message[i] + "\n"
    else:
        txt = message[int(click_event.value)]
    if len(txt) < 60:
        block1 = ImageBlock(
            image_url="https://mblogthumb-phinf.pstatic.net/MjAxODAxMTVfMjY3/MDAxNTE1OTk5MDY2MTI3.WAC3ZFTcc4LgCpCoHd2XRBUUyxQI2MLkQvVNceb1-yIg.ndOybGvKucvJEln42N_Cwv7OXpj7UtBM0V8Yi4mO6ukg.JPEG.dhkglatpek/1.jpg?type=w2",
            alt_text="비안와요"
        )
        gRain = "raining"
    elif len(txt) < 70:
        block1 = ImageBlock(
            image_url="https://previews.123rf.com/images/alekseyvanin/alekseyvanin1606/alekseyvanin160600085/59593058-%EB%B9%84-%EC%95%84%EC%9D%B4%EC%BD%98.jpg",
            alt_text="비와요"
        )
        gRain = "nraining"
    else:
        slack_web_client.chat_postMessage(
            channel=click_event.channel.id,
            text="*3일 뒤부터의 날씨는 다음과 같습니다. :*\n" + txt
        )
        return "OK", 200

    block2 = SectionBlock(
        text="*서울의 날씨는 다음과 같습니다. :*\n" + txt + "\n\n*여행할 목적지를 구 단위로 입력해 주세요. : * (ex: 강남구)",
        accessory=block1
    )

    my_blocks = [block2]
    slack_web_client.chat_postMessage(
        channel=click_event.channel.id,
        blocks=extract_json(my_blocks)
    )


# 챗봇이 멘션을 받았을 경우
@slack_events_adaptor.on("app_mention")
def app_mentioned(event_data):
    channel = event_data["event"]["channel"]
    text = event_data["event"]["text"]
    Thread(target=inputText, args=(channel, text)).start()

# 챗봇이 버튼 입력을 받았을 경우
@app.route("/click", methods=["GET", "POST"])
def on_button_click():
    payload = request.values["payload"]
    click_event = MessageInteractiveEvent(json.loads(payload))
    Thread(target=buttonText, args=(click_event, )).start()
    return "OK", 200

if __name__ == '__main__':
    app.run('127.0.0.1', port=8080)
